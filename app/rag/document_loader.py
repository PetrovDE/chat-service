# app/rag/document_loader.py
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from langchain_core.documents import Document
from langchain_community.document_loaders import (
    PyPDFLoader,
    Docx2txtLoader,
    TextLoader,
    CSVLoader,
    JSONLoader,
    UnstructuredExcelLoader,
    UnstructuredMarkdownLoader
)
import json
from app.rag.config import rag_config

logger = logging.getLogger(__name__)


class DocumentLoader:
    """
    Универсальный загрузчик документов
    Поддерживает: PDF, DOCX, TXT, CSV, Excel, JSON, Markdown
    """

    def __init__(self):
        """Инициализация загрузчика"""
        self.supported_loaders = {
            '.pdf': self._load_pdf,
            '.docx': self._load_docx,
            '.doc': self._load_docx,
            '.txt': self._load_text,
            '.csv': self._load_csv,
            '.xlsx': self._load_excel,
            '.xls': self._load_excel,
            '.json': self._load_json,
            '.md': self._load_markdown,
        }
        logger.info("✅ DocumentLoader initialized")

    def load_file(
            self,
            file_path: str,
            metadata: Optional[Dict[str, Any]] = None
    ) -> List[Document]:
        """
        Загрузить файл и вернуть список Document объектов

        Args:
            file_path: Путь к файлу
            metadata: Дополнительные метаданные

        Returns:
            Список Document объектов
        """
        try:
            path = Path(file_path)

            # Проверка существования файла
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            # Проверка размера файла
            file_size = path.stat().st_size
            if not rag_config.validate_file_size(file_size):
                raise ValueError(
                    f"File too large: {file_size / 1024 / 1024:.1f}MB "
                    f"(max: {rag_config.max_file_size_mb}MB)"
                )

            # Определить расширение
            extension = path.suffix.lower()

            # Проверка поддержки
            if extension not in self.supported_loaders:
                raise ValueError(f"Unsupported file type: {extension}")

            logger.info(f"📂 Loading file: {path.name} ({extension})")

            # Загрузить файл с помощью соответствующего loader
            loader_func = self.supported_loaders[extension]
            documents = loader_func(file_path)

            # Добавить метаданные
            if metadata:
                for doc in documents:
                    doc.metadata.update(metadata)

            # Добавить базовые метаданные
            for doc in documents:
                doc.metadata.update({
                    'source': str(path),
                    'file_name': path.name,
                    'file_type': extension,
                    'file_size': file_size
                })

            logger.info(f"✅ Loaded {len(documents)} documents from {path.name}")
            return documents

        except Exception as e:
            logger.error(f"❌ Error loading file {file_path}: {e}")
            raise

    def _load_pdf(self, file_path: str) -> List[Document]:
        """Загрузить PDF файл"""
        try:
            loader = PyPDFLoader(file_path)
            documents = loader.load()

            # Добавить номера страниц в метаданные
            for idx, doc in enumerate(documents):
                doc.metadata['page'] = idx + 1

            logger.debug(f"📕 Loaded PDF: {len(documents)} pages")
            return documents

        except Exception as e:
            logger.error(f"❌ Error loading PDF: {e}")
            # Fallback: попробовать прочитать как текст
            return self._load_as_text(file_path, "PDF (fallback)")

    def _load_docx(self, file_path: str) -> List[Document]:
        """Загрузить DOCX файл"""
        try:
            loader = Docx2txtLoader(file_path)
            documents = loader.load()
            logger.debug(f"📄 Loaded DOCX: {len(documents)} documents")
            return documents

        except Exception as e:
            logger.error(f"❌ Error loading DOCX: {e}")
            return self._load_as_text(file_path, "DOCX (fallback)")

    def _load_text(self, file_path: str) -> List[Document]:
        """Загрузить TXT файл"""
        try:
            loader = TextLoader(file_path, encoding='utf-8')
            documents = loader.load()
            logger.debug(f"📝 Loaded TXT: {len(documents)} documents")
            return documents

        except UnicodeDecodeError:
            # Попробовать другие кодировки
            for encoding in ['cp1251', 'latin-1', 'iso-8859-1']:
                try:
                    loader = TextLoader(file_path, encoding=encoding)
                    documents = loader.load()
                    logger.debug(f"📝 Loaded TXT with {encoding}: {len(documents)} documents")
                    return documents
                except:
                    continue

            logger.error("❌ Failed to decode text file with any encoding")
            raise ValueError("Unable to decode text file")

    def _load_csv(self, file_path: str) -> List[Document]:
        """Загрузить CSV файл"""
        try:
            loader = CSVLoader(
                file_path=file_path,
                encoding='utf-8',
                csv_args={'delimiter': ','}
            )
            documents = loader.load()

            # Добавить номера строк
            for idx, doc in enumerate(documents):
                doc.metadata['row'] = idx + 1

            logger.debug(f"📊 Loaded CSV: {len(documents)} rows")
            return documents

        except Exception as e:
            logger.error(f"❌ Error loading CSV: {e}")
            # Попробовать с другим разделителем
            try:
                loader = CSVLoader(
                    file_path=file_path,
                    encoding='utf-8',
                    csv_args={'delimiter': ';'}
                )
                documents = loader.load()
                logger.debug(f"📊 Loaded CSV with semicolon: {len(documents)} rows")
                return documents
            except:
                return self._load_as_text(file_path, "CSV (fallback)")

    def _load_excel(self, file_path: str) -> List[Document]:
        """Загрузить Excel файл"""
        try:
            loader = UnstructuredExcelLoader(file_path, mode="elements")
            documents = loader.load()
            logger.debug(f"📈 Loaded Excel: {len(documents)} elements")
            return documents

        except Exception as e:
            logger.error(f"❌ Error loading Excel: {e}")
            # Fallback: попробовать прочитать с openpyxl напрямую
            return self._load_excel_fallback(file_path)

    def _load_excel_fallback(self, file_path: str) -> List[Document]:
        """Альтернативная загрузка Excel с openpyxl"""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            documents = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Прочитать все строки
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                    if row_text.strip():
                        rows.append(row_text)

                # Создать документ для каждого листа
                if rows:
                    content = '\n'.join(rows)
                    doc = Document(
                        page_content=content,
                        metadata={
                            'sheet_name': sheet_name,
                            'rows_count': len(rows)
                        }
                    )
                    documents.append(doc)

            logger.debug(f"📈 Loaded Excel (fallback): {len(documents)} sheets")
            return documents

        except Exception as e:
            logger.error(f"❌ Excel fallback failed: {e}")
            return []

    def _load_json(self, file_path: str) -> List[Document]:
        """Загрузить JSON файл"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # Преобразовать JSON в текст
            if isinstance(data, dict):
                content = json.dumps(data, indent=2, ensure_ascii=False)
            elif isinstance(data, list):
                content = json.dumps(data, indent=2, ensure_ascii=False)
            else:
                content = str(data)

            doc = Document(
                page_content=content,
                metadata={'json_type': type(data).__name__}
            )

            logger.debug(f"📋 Loaded JSON: 1 document")
            return [doc]

        except Exception as e:
            logger.error(f"❌ Error loading JSON: {e}")
            return self._load_as_text(file_path, "JSON (fallback)")

    def _load_markdown(self, file_path: str) -> List[Document]:
        """Загрузить Markdown файл"""
        try:
            loader = UnstructuredMarkdownLoader(file_path)
            documents = loader.load()
            logger.debug(f"📝 Loaded Markdown: {len(documents)} documents")
            return documents

        except Exception as e:
            logger.error(f"❌ Error loading Markdown: {e}")
            return self._load_as_text(file_path, "Markdown (fallback)")

    def _load_as_text(self, file_path: str, source_type: str) -> List[Document]:
        """
        Fallback: загрузить файл как обычный текст
        Используется когда специализированный loader не работает
        """
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()

            doc = Document(
                page_content=content,
                metadata={'loaded_as': source_type}
            )

            logger.warning(f"⚠️ Loaded as text fallback ({source_type})")
            return [doc]

        except Exception as e:
            logger.error(f"❌ Text fallback failed: {e}")
            return []

    def load_from_db_file(self, db_file) -> List[Document]:
        """
        Загрузить файл из записи БД

        Args:
            db_file: Объект File из БД (из app.database.models)

        Returns:
            Список Document объектов
        """
        try:
            metadata = {
                'file_id': str(db_file.id),
                'user_id': str(db_file.user_id) if db_file.user_id else None,
                'original_filename': db_file.original_filename,
                'uploaded_at': str(db_file.created_at)
            }

            # Если есть full_content, использовать его
            if db_file.full_content:
                logger.info(f"📄 Using cached content from DB for {db_file.original_filename}")
                doc = Document(
                    page_content=db_file.full_content,
                    metadata=metadata
                )
                return [doc]

            # Иначе загрузить из файла
            return self.load_file(db_file.file_path, metadata)

        except Exception as e:
            logger.error(f"❌ Error loading from DB file: {e}")
            raise

    def get_supported_extensions(self) -> List[str]:
        """Получить список поддерживаемых расширений"""
        return list(self.supported_loaders.keys())

    def is_supported(self, filename: str) -> bool:
        """Проверить, поддерживается ли файл"""
        extension = Path(filename).suffix.lower()
        return extension in self.supported_loaders